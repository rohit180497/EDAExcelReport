import pandas as pd
import numpy as np
import datetime
from sys import maxsize
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

class EDA_Formatter:
  
    def __init__(self, path= "..\\reports\\EDA_raw.xlsx", model_type= "Target", conditional_color ="red"):


        """
        A class to format and generate an Excel report for EDA results with conditional formatting.

        Parameters
        ----------
        path : str, optional
            The path to the raw EDA Excel report (default is "..\\reports\\EDA_raw.xlsx").
        model_type : str, optional
            The type of model used for generating the EDA (default is "Target").
        conditional_color : str, optional
            The color used for conditional formatting in the report (default is "red").
        """

        """
        Initializes the EDA_Formatter with the given parameters and runs the formatter.
        """
        self.input_path = path
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

        self.output_path = f"_{timestamp}.xlsx".join(path.split(".xlsx"))
        self.type = model_type
        self.color = conditional_color
        self.r = 1
        self.c = 1
        
        self.setup_workbook()
        self.run_formatter()

    def setup_workbook(self):

        """
        Sets up the workbook by loading the initial sheet and setting column widths.
        """
        self.wb = Workbook()
        ws2 = load_workbook(self.input_path)['ROC Report']
        ws2._parent = self.wb
        self.wb._add_sheet(ws2)                                                                     
        self.ws = self.wb.worksheets[0]
        self.ws_title = 'Detailed EDA'
        self.set_column_widths()

    def run_formatter(self):
        """
        Runs the formatter to process and format the EDA results.
        """
        df = pd.read_excel(self.input_path, "Detailed EDA", engine = "openpyxl")
        df.rename(columns = {"value": "Value",
                             "count": "Frequency",
                             "sum": self.type,
                             "mean": self.type+" Rate"}, inplace=True)
        df.insert(loc=3, column="Freq Distribution", value=0)
        df.insert(loc=6, column= r"% of Total "+ self.type, value=0)
        df.insert(loc=7, column="Lift", value = 0)


        df2 = pd.DataFrame(columns=df.columns)
        #init_val = df.loc[0][0]
        init_val = df.iloc[0,0]

        for index in df.index:

            if df.iloc[index, 0] != init_val:
                df2.reset_index(inplace=True, drop=True)
                self.write_to_excel(df2)
                self.r += len(df2.index) + 3

                df2 = pd.DataFrame(columns= df.columns)
                init_val= df.iloc[index, 0]

            df2.loc[index] = df.loc[index]
        df2.reset_index(inplace=True, drop=True)
        self.write_to_excel(df2)

        self.wb.save(self.output_path)
        print(f"Your EDA report is ready at {self.output_path}")


    @staticmethod
    def is_number(n):

        """
        Checks if the given value is a number.

        Parameters
        ----------
        n : Any
            The value to check.

        Returns
        -------
        bool
            True if the value is a number, False otherwise.
        """
        try:
            float(n)
        except ValueError:
            return False
        return True
    
    @staticmethod
    def sort(df):
        """
        Sorts the DataFrame by the 'Value' column.

        Parameters
        ----------
        df : pd.DataFrame
            The DataFrame to sort.

        Returns
        -------
        pd.DataFrame
            The sorted DataFrame.
        """

        df["Value"] = df["Value"].apply(pd.to_numeric)
        df.sort_values(by = "Value", inplace=True)
        df = df.reset_index(drop=True)

        return df

    def set_column_widths(self):
        """
        Sets the column widths for the worksheet.
        """
         
        width_list = [43, 50, 10, 15, 10, 14, 19, 11]

        for index, width in enumerate(width_list):
            char = get_column_letter(index+self.c)
            self.ws.column_dimensions[char].width = (width+0.78)

    def number_format(self, df_rows):

        """
        Formats numeric values in the worksheet cells.

        Parameters
        ----------
        df_rows : int
            The number of rows in the DataFrame.
        """

        char = get_column_letter(self.c+1)

        prev_cell = self.ws[char + str(self.r)]
        prev_val = prev_cell.value

        for row in range(1, df_rows+1):

            cell = self.ws[char+str(self.r + row)]
            val = cell.value

            if (row==1):
                cell.value = "<= " + str(val)
            elif (row == df_rows):
                cell.value = "> " + str(prev_val)
            else:
                cell.value = "> "+str(prev_val)+ " & <= "+str(val) 

            prev_val = val


    def df_formatter(self, rows, cols):

        """
        Applies formatting to the DataFrame cells in the worksheet.

        Parameters
        ----------
        rows : int
            The number of rows in the DataFrame.
        cols : int
            The number of columns in the DataFrame.
        """


        thin_border = Border(left=Side(border_style='thin'),
                             right=Side(border_style='thin'),
                             top=Side(border_style='thin'),
                             bottom=Side(border_style='thin'))
        font = Font(name='calibri',
                    size= 11,
                    bold=True)
        fill = PatternFill(fill_type='solid',
                           start_color='E4DFEC',
                           end_color='E4DFEC')
        center_alignment = Alignment(
            horizontal='center', vertical='center')
        left_alignment = Alignment(
            horizontal='center', vertical='center')
        right_alignment = Alignment(
            horizontal='center', vertical='center')
        
        for col in range(0, cols):

            char = get_column_letter(col+self.c)
            header_cell = self.ws[char+str(self.r)]

            header_cell.border = thin_border
            header_cell.font = font
            header_cell.alignment = center_alignment
            header_cell.fill = fill

            for row in range(0, rows):
                data_cell = self.ws[char+ str(self.r+row+1)]
                if col == 0:
                    data_cell.alignment = center_alignment
                else:
                    if (col==3 or col == 5 or col== 6):
                        data_cell.number_format = '0.00%'
                    elif (col == 7):
                        data_cell.number_format = '0.00'
                    if col == 1:
                        data_cell.alignment = left_alignment
                    else:
                        data_cell.alignment =  right_alignment
            
        char = get_column_letter(self.c)
        self.ws.merge_cells(char+str(self.r+1)+':'+char+str(self.r+rows))

    def cond_format(self, start_row, end_row, start_col, end_col):
        """Apply conditional formatting to the worksheet."""

        if (self.color == 'red' or self.color == 'Red'):
            high_value = 'F8696B'
        else:
            high_value = '63BE7B'


        for col in range(start_col, end_col+1):
            char = get_column_letter(col)
            
            if (self.color == 'color' or self.color == 'Color'):
                self.ws.conditional_formatting.add(char+str(start_row)+ ':'+char+str(end_row),
                                                   ColorScaleRule(start_type='percentile', start_value=0, start_color='F8696B',
                                                                  mid_type='percentile', mid_value= 50, mid_color= 'FFEB84',
                                                                  end_type= 'percentile', end_value= 100, end_color= '63BE7B'))
            else:
                self.ws.conditional_formatting.add(char+str(start_row)+':'+char+str(end_row),
                                                   ColorScaleRule(start_type='percentile', start_value= 0, start_color= 'FCFCFF',
                                                                  end_type= 'percentile', end_value=100, end_color= high_value)
                                                   )
                

    def write_to_excel(self, df):
        
        """Write the DataFrame to the worksheet and apply formatting."""

        num_flag = False

        thin_border = Border(bottom=Side(style='thin'))


        #writing header
        for col in range(0, len(df.columns)):
            char = get_column_letter(self.c+ col)
            cell = self.ws[char+str(self.r)]
            cell.value  = df.columns[col]

        #sorting if numeric after replacing inf with maxsize

        if (self.is_number(df.iloc[0, 1]) or df.iloc[0, 1]== 'inf'):
            for index in df.index:
                if df.iloc[index, 1]== 'inf':
                    df.at[index, "Value"]= maxsize
            df = self.sort(df)
            df['Value'] = df['Value'].round(decimals= 2)
            num_flag = True


        for row in range(0, len(df.index)):
            for col in range(0, len(df.columns)):
                char = get_column_letter(self.c+col) 
                cell = self.ws[char+str(self.r+ row+1)]
                if (col== 3):
                    char2 = get_column_letter(self.c + 2)
                    cell.value = "="+char2+str(self.r+row+1)+'/SUM('+char2+ '$' + str(self.r +1)+':'+char2+ '$'+str(self.r+len(df.index))+')'
                elif (col== 6):
                    char2 = get_column_letter(self.c + 4)
                    cell.value = "="+char2+str(self.r+row+1)+'/SUM('+char2+ '$' + str(self.r +1)+':'+char2+ '$'+str(self.r+len(df.index))+')'
                elif (col== 7):
                    char2 = get_column_letter(self.c+3)
                    char3 = get_column_letter(self.c+6)
                    cell.value = "="+char3 + \
                        str(self.r+row+1)+ '/'+ char2+ str(self.r+row+1)
                else:
                    cell.value = df.iloc[row, col]

        # Adding a border to the last row of the data
        last_row = self.r + len(df.index)
        for col in range(0, len(df.columns)):
            char = get_column_letter(self.c + col)
            cell = self.ws[char + str(last_row)]
            cell.border = thin_border

        if num_flag:
            self.number_format(len(df.index))

        self.df_formatter(rows= len(df.index), cols= len(df.columns))
        self.cond_format(start_row= self.r+1,
                         end_row= self.r+ len(df.index), start_col = self.c + 5, end_col=self.c+ 7)
        
        self.ws.sheet_view.showGridLines = False



    def __repr__(self):
        return "<EDAReport: Your report is ready>"
        









